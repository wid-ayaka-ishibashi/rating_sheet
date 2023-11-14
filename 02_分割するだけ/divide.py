# %%
import os
import openpyxl as px
from IPython.display import display

divide_pre = "01_分割前"
divide_after = "02_分割後"
save_file_name1 = "2023年上期評価シート("

for folder, subFolder, files in os.walk(divide_pre):
    for file in files:
#        ext = os.path.splitext(file)[0] #拡張子変更考慮
        if "_評価済み" in file:
            wb = px.load_workbook(folder + "\\" + file)
            i = 0
            for sheetname in wb.sheetnames:
                if "【" and "】" not in sheetname:
                    i = i+1
                try:
                    ws_target = wb.worksheets[i].title
                except:
                    break
                    ws = wb[sheetname]
                    peple_fullname = ws["A2"].value
                    for sheetname in wb.sheetnames:
                        #[ws_target]以外は削除
                        if sheetname.title == ws_target:
                            continue
                        else:
                            wb.remove(sheetname)
                        new_wb.save(divide_after + "\\" + save_file_name1 + peple_fullname + ").xlsx")

# %%
